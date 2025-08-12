import os
import logging
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def salvar_xlsx_atomic(path, df, options=None):
    tmp_file = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="sync_") as tmp:
            tmp_file = tmp.name

        with pd.ExcelWriter(tmp_file, engine='openpyxl', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
            df.to_excel(writer, index=False, sheet_name='data')

        wb = load_workbook(tmp_file)
        ws = wb['data']

        opts = options or {}
        force_text = set(opts.get('force_text', []))
        force_integer = set(opts.get('force_integer', []))
        force_date = set(opts.get('force_date', []))

        headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

        for header, col_idx in headers.items():
            if header is None:
                continue
            
            col_letter = get_column_letter(col_idx)
            fmt = 'General'
            
            if header in force_text:
                fmt = '@'
            elif header in force_integer:
                fmt = '0'
            elif header in force_date:
                fmt = 'yyyy-mm-dd;@'
                
            for cell in ws[col_letter][1:]:
                cell.number_format = fmt
        
        wb.save(tmp_file)

        os.replace(tmp_file, path)
        logging.info(f"Arquivo salvo: {path} (linhas={len(df)})")

    except Exception as e:
        logging.exception(f"Falha ao salvar ou formatar o arquivo XLSX '{path}': {e}")
    finally:
        if tmp_file and os.path.exists(tmp_file):
            try:
                if not os.path.exists(path) or os.path.getmtime(tmp_file) > os.path.getmtime(path):
                    os.remove(tmp_file)
            except OSError:
                pass

def salvar_xlsx_em_chunks_atomic(path, df_chunks, options=None):
    """
    Salva um arquivo .xlsx de forma atômica a partir de um iterador de DataFrames (chunks).
    """
    tmp_file = None
    total_rows = 0
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="sync_") as tmp:
            tmp_file = tmp.name

        with pd.ExcelWriter(tmp_file, engine='openpyxl', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
            header = True
            for df_chunk in df_chunks:
                df_chunk.to_excel(writer, index=False, sheet_name='data', header=header, startrow=writer.sheets['data'].max_row if not header else 0)
                total_rows += len(df_chunk)
                header = False # O cabeçalho é escrito apenas uma vez

        # A formatação de colunas é mais complexa em modo de chunk e pode ser omitida
        # para grandes arquivos para economizar memória, ou aplicada de forma simplificada.
        # Por simplicidade, vamos omitir a formatação detalhada por coluna nesta versão.

        os.replace(tmp_file, path)
        logging.info(f"Arquivo salvo: {path} (total de linhas={total_rows})")

    except Exception as e:
        logging.exception(f"Falha ao salvar ou formatar o arquivo XLSX em chunks '{path}': {e}")
    finally:
        if tmp_file and os.path.exists(tmp_file):
            try:
                os.remove(tmp_file)
            except OSError:
                pass